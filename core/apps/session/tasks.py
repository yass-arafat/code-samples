import logging
import uuid
from datetime import date, datetime, timedelta

import requests
from celery import shared_task
from django.conf import settings
from openpyxl import load_workbook

from config import emails
from core.apps.common.common_functions import get_timezone_offset_from_datetime_diff
from core.apps.common.const import (
    FIRST_TIMEZ0NE_OFFSET,
    MIDNIGHT_CRONJOB_TIME,
    MORNING_CRONJOB_TIME,
    TIME_RANGE_BOUNDARY,
)
from core.apps.common.date_time_utils import (
    convert_str_date_time_to_date_time_obj,
    convert_timezone_offset_to_seconds,
    time_diff_between_two_timezone_offsets,
)
from core.apps.common.models import CronHistoryLog
from core.apps.evaluation.daily_evaluation.utils import (
    day_midnight_calculation,
    day_morning_calculation,
)
from core.apps.user_auth.models import UserAuthModel
from core.apps.user_profile.models import UserActivityLog

from ..activities.enums import SecondBySecondDataEnum
from ..activities.utils import dakghor_get_athlete_activity
from ..common.enums.service_enum import ServiceType
from ..common.tp_common_utils import (
    read_s3_pillar_heart_rate_data,
    read_s3_pillar_power_data,
)
from ..common.utils import download_s3_file, log_extra_fields, upload_file_to_s3
from ..evaluation.session_evaluation.dictionary import (
    get_hr_data_dict_with_threshold,
    get_power_data_dict_with_threshold,
)
from ..user_profile.utils import get_user_fthr, get_user_ftp, get_user_max_heart_rate
from .models import ActualSession
from .utils import flag_unusual_load_sqs_drop

logger = logging.getLogger(__name__)


def get_cron_name(cron_code_value):
    cron_codes = CronHistoryLog.CronCode
    for cron_code in cron_codes:
        if cron_code == cron_code_value:
            return cron_code.label
    return ""


def is_cron_already_done_for_user(
    user, cron_start_time, first_cron_start_time, cron_code
):
    user_last_cron = CronHistoryLog.objects.filter(
        cron_code=cron_code, user_auth=user
    ).last()

    if (
        user_last_cron
        and first_cron_start_time
        <= user_last_cron.timestamp.replace(tzinfo=None)
        <= cron_start_time
    ):
        activity_codes = (
            UserActivityLog.ActivityCode.SUCCESSFUL_VALID_GARMIN_ACTIVITY_FILE_CALCULATION.value,
            UserActivityLog.ActivityCode.SUCCESSFUL_VALID_STRAVA_DATA_CALCULATION.value,
        )
        user_last_activity = (
            UserActivityLog.objects.filter(
                user_auth=user, activity_code__in=activity_codes
            )
            .order_by("timestamp")
            .last()
        )

        if (
            user_last_activity
            and user_last_activity.timestamp > user_last_cron.timestamp
        ):
            return False

        return True
    return False


def midnight_calculation():
    try:
        # If it takes 30 seconds or more to get here the calculation will be faulty
        start_time = datetime.now().replace(second=0, microsecond=0)
        user_count = 0
        error_msg = ""

        timezone_offset = get_timezone_offset_from_datetime_diff(
            datetime.combine(date.today(), MIDNIGHT_CRONJOB_TIME) - start_time
        )
        first_cron_time_diff = time_diff_between_two_timezone_offsets(
            FIRST_TIMEZ0NE_OFFSET, timezone_offset
        )
        first_cron_start_time = start_time - timedelta(seconds=first_cron_time_diff)

        user_local_date = (
            start_time
            + timedelta(seconds=convert_timezone_offset_to_seconds(timezone_offset))
        ).date()

        cron_code = CronHistoryLog.CronCode.MIDNIGHT_CALCULATION
        cron_logs = []

        users = UserAuthModel.objects.filter(
            is_active=True,
            profile_data__timezone__offset=timezone_offset,
            profile_data__is_active=True,
        )

        flagged_user_list = []
        try:
            for user in users:
                if is_cron_already_done_for_user(
                    user, start_time, first_cron_start_time, cron_code
                ):
                    continue
                try:
                    day_midnight_calculation(user, user_local_date)
                except Exception as e:
                    log_message = "usr_id = " + str(user.id) + "  Error: " + str(e)
                    logger.error("In midnight calculation " + log_message)
                    error_msg += str(e) + "\n"
                    cron_status = CronHistoryLog.StatusCode.FAILED
                else:
                    user_count += 1
                    cron_status = CronHistoryLog.StatusCode.SUCCESSFUL

                cron_logs.append(
                    CronHistoryLog(
                        cron_code=cron_code,
                        user_auth=user,
                        user_id=user.code,
                        status=cron_status,
                    )
                )

                if flag_unusual_load_sqs_drop(user, user_local_date):
                    flagged_user_list.append(user.id)

            CronHistoryLog.objects.bulk_create(cron_logs)

        except Exception as e:
            logger.error(
                f"Midnight cron failed for timezone {timezone_offset}\nException: {str(e)}"
            )
            flagged_user_list = []
            for user in users:
                if flag_unusual_load_sqs_drop(user, user_local_date):
                    flagged_user_list.append(user.id)

        if flagged_user_list:
            logger.info(
                f"Actual load or SQS changed more than change limit for {[user for user in flagged_user_list]} \
                           at {user_local_date} when midnight cron failed."
            )
            emails.send_unusual_load_drop_email(
                flagged_user_list, "Midnight", timezone_offset
            )

        end_time = datetime.now()
        total_time = (end_time - start_time).total_seconds() * 1000000
        emails.send_midnight_calculation_email_notification(
            user_count, start_time, total_time, error_msg
        )
    except Exception as e:
        logger.error(f"Midnight Job failed. Exception: {e}")
        requests.post(
            url=settings.ASYNC_HEALTHCHECK_URL + "/fail",
            data={"Exception": e, "Cron": "Midnight Calculation"},
        )


def morning_calculation():
    try:
        start_time = datetime.now().replace(second=0, microsecond=0)
        user_count = 0
        error_msg = ""

        timezone_offset = get_timezone_offset_from_datetime_diff(
            datetime.combine(date.today(), MORNING_CRONJOB_TIME) - start_time
        )
        first_cron_time_diff = time_diff_between_two_timezone_offsets(
            FIRST_TIMEZ0NE_OFFSET, timezone_offset
        )
        first_cron_start_time = start_time - timedelta(seconds=first_cron_time_diff)

        user_local_date = (
            start_time
            + timedelta(
                days=1, seconds=convert_timezone_offset_to_seconds(timezone_offset)
            )
        ).date()

        cron_code = CronHistoryLog.CronCode.MORNING_CALCULATION
        cron_logs = []

        users = UserAuthModel.objects.filter(
            is_active=True,
            profile_data__timezone__offset=timezone_offset,
            profile_data__is_active=True,
        )

        flagged_user_list = []
        try:
            for user in users:
                # write_log(f"Starting loop for user: {user.id}", file_name)
                if is_cron_already_done_for_user(
                    user, start_time, first_cron_start_time, cron_code
                ):
                    continue

                try:
                    day_morning_calculation(user, user_local_date)
                except Exception as e:
                    error_msg += str(e) + "\n"
                    cron_status = CronHistoryLog.StatusCode.FAILED
                else:
                    user_count += 1
                    cron_status = CronHistoryLog.StatusCode.SUCCESSFUL

                cron_logs.append(
                    CronHistoryLog(
                        cron_code=cron_code,
                        user_auth=user,
                        user_id=user.code,
                        status=cron_status,
                    )
                )

                if flag_unusual_load_sqs_drop(user, user_local_date):
                    flagged_user_list.append(user.id)

            CronHistoryLog.objects.bulk_create(cron_logs)

        except Exception as e:
            logger.error(f"Morning cron failed for timezone {timezone_offset} {str(e)}")

            flagged_user_list = []
            for user in users:
                if flag_unusual_load_sqs_drop(user, user_local_date):
                    flagged_user_list.append(user.id)

        if flagged_user_list:
            logger.info(
                f"Actual load or SQS changed more than change limit for {[user for user in flagged_user_list]} \
                           at {user_local_date} when morning cron failed."
            )
            emails.send_unusual_load_drop_email(
                flagged_user_list, "Morning", timezone_offset
            )

        end_time = datetime.now()
        total_time = (end_time - start_time).total_seconds() * 1000000
        emails.send_morning_calculation_email_notification(
            user_count, start_time, total_time, error_msg
        )

    except Exception as e:
        logger.exception(f"Morning Job failed. Exception: {e}")
        requests.post(
            url=settings.ASYNC_HEALTHCHECK_URL + "/fail",
            data={"Exception": e, "Cron": "Morning Calculation"},
        )


@shared_task
def populate_compressed_power_hr_data(
    user_auth_id, athlete_activity_code, session_date_time
):
    try:
        logger.info(
            f"Compressing power and hr data of Activity: {athlete_activity_code}"
        )
        session_date_time = convert_str_date_time_to_date_time_obj(session_date_time)
        user_auth = UserAuthModel.objects.get(id=user_auth_id)

        logger.info("Fetching data from dakghor")
        s3_pillar_file_path = dakghor_get_athlete_activity(
            athlete_activity_code
        ).json()["data"]["athlete_activity"]["pillar_file"]

        logger.info("Downloading file from S3")
        temporary_file_path = download_s3_file(s3_pillar_file_path)

        workbook = load_workbook(temporary_file_path)
        worksheet = workbook.active

        logger.info("Processing power data")
        user_ftp = get_user_ftp(user_auth, session_date_time)
        power_data = read_s3_pillar_power_data(worksheet)
        compressed_power_data = get_power_data_dict_with_threshold(
            power_data, user_ftp, total_point=250
        )
        row = 2
        worksheet.cell(
            row=1, column=SecondBySecondDataEnum.POWER_250.value[0] + 1
        ).value = SecondBySecondDataEnum.POWER_250.value[1]
        worksheet.cell(
            row=1, column=SecondBySecondDataEnum.POWER_250_TIME.value[0] + 1
        ).value = SecondBySecondDataEnum.POWER_250_TIME.value[1]
        worksheet.cell(
            row=1, column=SecondBySecondDataEnum.POWER_250_ZONE.value[0] + 1
        ).value = SecondBySecondDataEnum.POWER_250_ZONE.value[1]
        for power in compressed_power_data:
            worksheet.cell(
                row=row, column=SecondBySecondDataEnum.POWER_250.value[0] + 1
            ).value = power["value"]
            worksheet.cell(
                row=row, column=SecondBySecondDataEnum.POWER_250_TIME.value[0] + 1
            ).value = power["time"]
            worksheet.cell(
                row=row, column=SecondBySecondDataEnum.POWER_250_ZONE.value[0] + 1
            ).value = power["zone_focus"]
            row += 1

        logger.info("Processing heart rate data")
        user_fthr = get_user_fthr(user_auth, session_date_time)
        user_max_hr = get_user_max_heart_rate(user_auth, session_date_time)
        hr_data = read_s3_pillar_heart_rate_data(worksheet)
        compressed_hr_data = get_hr_data_dict_with_threshold(
            hr_data, user_fthr, user_max_hr, total_point=250
        )
        row = 2
        worksheet.cell(
            row=1, column=SecondBySecondDataEnum.HEART_RATE_250.value[0] + 1
        ).value = SecondBySecondDataEnum.HEART_RATE_250.value[1]
        worksheet.cell(
            row=1, column=SecondBySecondDataEnum.HEART_RATE_250_TIME.value[0] + 1
        ).value = SecondBySecondDataEnum.HEART_RATE_250_TIME.value[1]
        worksheet.cell(
            row=1, column=SecondBySecondDataEnum.HEART_RATE_250_ZONE.value[0] + 1
        ).value = SecondBySecondDataEnum.HEART_RATE_250_ZONE.value[1]
        for hr in compressed_hr_data:
            worksheet.cell(
                row=row, column=SecondBySecondDataEnum.HEART_RATE_250.value[0] + 1
            ).value = hr["value"]
            worksheet.cell(
                row=row, column=SecondBySecondDataEnum.HEART_RATE_250_TIME.value[0] + 1
            ).value = hr["time"]
            worksheet.cell(
                row=row, column=SecondBySecondDataEnum.HEART_RATE_250_ZONE.value[0] + 1
            ).value = hr["zone_focus"]
            row += 1

        workbook.save(temporary_file_path)

        logger.info("Uploading compressed data file to S3")
        upload_file_to_s3(temporary_file_path, s3_pillar_file_path)

        logger.info(
            f"Successfully processed compressing power and hr data of Activity: {athlete_activity_code}"
        )
    except Exception as e:
        logger.exception(
            "Failed to populate compressed power hr data",
            extra=log_extra_fields(
                user_auth_id=user_auth_id,
                service_type=ServiceType.INTERNAL.value,
                exception_message=str(e),
            ),
        )


@shared_task
def populate_session_actual_intervals(user_auth_id, actual_session_id):
    from .services import SessionIntervalService

    try:
        user_auth = UserAuthModel.objects.get(id=user_auth_id)
        actual_session = ActualSession.objects.get(id=actual_session_id)
        actual_session = SessionIntervalService(
            user_auth, actual_session
        ).set_actual_intervals()
        if actual_session.actual_intervals:
            logger.info("Saving intervals")
            actual_session.save(update_fields=["actual_intervals"])
    except Exception as e:
        logger.exception(
            "Failed to populate session actual intervals",
            extra=log_extra_fields(
                user_auth_id=user_auth_id,
                service_type=ServiceType.INTERNAL.value,
                exception_message=str(e),
            ),
        )


@shared_task
def populate_user_actual_intervals(user_auth_id):
    try:
        actual_sessions = ActualSession.objects.filter(
            user_auth_id=user_auth_id,
            athlete_activity_code__isnull=False,
            session_code__isnull=False,
            actual_intervals__isnull=True,
            is_active=True,
        ).values("id")
        for actual_session in actual_sessions:
            populate_session_actual_intervals.delay(user_auth_id, actual_session["id"])
    except Exception as e:
        logger.exception(
            "Failed to populate user session actual intervals",
            extra=log_extra_fields(
                user_auth_id=user_auth_id,
                service_type=ServiceType.INTERNAL.value,
                exception_message=str(e),
            ),
        )


def populate_actual_session_code_admin_task(user):
    """Populate empty/null actual session code column for actual sessions of user (Admin task)"""
    actual_sessions = ActualSession.objects.filter(user_auth=user, is_active=True)
    for actual_session in actual_sessions:
        if actual_session.code is None:
            code = uuid.uuid4()
            ActualSession.objects.filter(pk=actual_session.pk).update(code=code)
        else:
            start_time = actual_session.session_date_time - timedelta(
                seconds=TIME_RANGE_BOUNDARY
            )
            end_time = actual_session.session_date_time + timedelta(
                seconds=TIME_RANGE_BOUNDARY
            )
            same_sessions = ActualSession.objects.filter(
                user_auth=actual_session.user_auth,
                session_date_time__range=(start_time, end_time),
                is_active=True,
            ).exclude(id=actual_session.id)
            for same_session in same_sessions:
                if same_session.code == actual_session.code:
                    code = uuid.uuid4()
                    ActualSession.objects.filter(pk=same_session.pk).update(code=code)
